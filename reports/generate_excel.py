import pandas as pd
from openpyxl import Workbook
from config import EXPORTS_DIR, PACKAGE_LEADS

class LeadReportExcel:
    def __init__(self, df, package):
        self.package = package
        self.leads = df.head(PACKAGE_LEADS[package]) if package in PACKAGE_LEADS else df.head(50)

    def generate(self, filename=None):
        if filename is None:
            filename = f"{EXPORTS_DIR}{self.package}_leads_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Business Leads"

        # Write headers
        headers = ['ID', 'Business Name', 'Phone', 'Address', 'Category']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, (_, row) in enumerate(self.leads.iterrows(), 2):
            ws.cell(row=row_num, column=1, value=row.get('id', 'N/A'))
            ws.cell(row=row_num, column=2, value=row.get('name', 'N/A'))
            ws.cell(row=row_num, column=3, value=row.get('phone', 'N/A'))
            ws.cell(row=row_num, column=4, value=row.get('address', 'N/A'))
            ws.cell(row=row_num, column=5, value=row.get('category', 'N/A'))

        wb.save(filename)
        return filename

if __name__ == '__main__':
    # Test with sample data
    sample_df = pd.DataFrame([
        {'id': 1, 'name': 'ABC Realty', 'phone': '6 12 345 678', 'address': 'Main St, Maseru', 'category': 'real_estate'},
        {'id': 2, 'name': 'DEF Agents', 'phone': '6 23 498 765', 'address': 'Highway Rd, Qoaling', 'category': 'real_estate'},
    ])
    generator = LeadReportExcel(sample_df, 'basic')
    excel_file = generator.generate('test_leads.xlsx')
    print(f"Excel generated: {excel_file}")
